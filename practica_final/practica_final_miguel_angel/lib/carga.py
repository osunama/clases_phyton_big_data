import csv
from openpyxl import load_workbook
import json
import xml.etree.ElementTree as et

def cargar_csv(carpeta, fichero):
    fichero = open(f"{carpeta}/{fichero}", "r", encoding='UTF-8')
    lector = csv.DictReader(fichero) 
    lista_soundwave = list(lector)
    
    fichero.close()
    return(lista_soundwave)

def cargar_excel(carpeta, fichero):
    excel = load_workbook(f'./{carpeta}/{fichero}')
    hoja = excel.active
    lista_artistas = []

    filas = hoja.iter_rows(values_only = True)
    cabeceras = next(filas)

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        empleado_dict = dict(zip(cabeceras, fila))
        lista_artistas.append(empleado_dict)
    return(lista_artistas)

def cargar_json(carpeta, nombre):
    try:
        fichero = open(f"./{carpeta}/{nombre}", "r", encoding="UTF-8")
        datos = json.load(fichero)
        return(datos)
    except FileNotFoundError:
        return('Archivo o carpeta no encontrado')

def cargar_xml(carpeta, fichero):
    fichero = et.parse(f'./{carpeta}/{fichero}')
    nodo_raiz = fichero.getroot()

    lista_patrocinadores = []

    for pat in nodo_raiz.findall('patrocinador'):
        patrocinador = {
            'nombre_empresa': pat.find('nombre_empresa').text,
            'contacto': pat.find('contacto').text,
            'email': pat.find('email').text,
            'importe_patrocinio': pat.find('importe_patrocinio').text,
            'categoria': pat.find('categoria').text,
            'fecha_inicio': pat.find('fecha_inicio').text,
            'fecha_fin': pat.find('fecha_fin').text,            
        }
        lista_patrocinadores.append(patrocinador)
    print (lista_patrocinadores)

  



    



