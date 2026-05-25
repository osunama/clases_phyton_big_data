from openpyxl import load_workbook, Workbook
import os
from lib.functions import limpiar_id, limpiar_texto, limpiar_precio, limpiar_stock

def cargar_excel(carpeta, archivo):
   excel = load_workbook(f"./{carpeta}/{archivo}")
   hoja = excel.active
   hoja = excel['Inventario']
   
   #cargamos las cabeceras
   filas = hoja.iter_rows(values_only = True)
   cabeceras = next(filas)
   
   lista_resultante = []
   for fila in hoja.iter_rows(min_row=2, values_only=True):
       producto = dict(zip(cabeceras, fila))
       lista_resultante.append(producto)
       
   return lista_resultante


def crear_excel(carpeta, archivo, datos):
    os.makedirs(carpeta, exist_ok=True)
    ruta = f"./{carpeta}/{archivo}"
    
    ## creamos el excel
    excel = Workbook()
    ## creamos la hoja
    hoja = excel.active
    hoja.title = 'Inventario_limpio'
    
    # Extraemos las cabeceras de un diccionario el primero por ejemplo
    cabeceras = list(datos[0].keys())
    hoja.append(cabeceras)
    
    for producto in datos:
        fila = [producto[clave] for clave in producto]
        hoja.append(fila)
    
    excel.save(ruta)
    print('Excel creado correctamente')
    
def procesar(lista):
    ## recorrer la lista y limpiar cada elemento de la lista. El objetivo de esta funcion es recibir una lista de datos sucia y devolverla limpia.
    lista_limpia = []
    for item in lista:
       item_limpio = {
           'id_producto': limpiar_id(item['id_producto']),
           'nombre': limpiar_texto(item['nombre']),
           'categoria': limpiar_texto(item['categoria'], True),
           'precio': limpiar_precio(item['precio']),
           'stock': limpiar_stock(item['stock'])
       }
       lista_limpia.append(item_limpio)
    return lista_limpia
    
    
    

## flujo cargar el fichero, limpiar y guardarlo
datos = cargar_excel('data', 'inventario_sucio.xlsx')
datos_limpios = procesar(datos)
crear_excel('data', 'inventario_limpio.xlsx', datos_limpios )
       