from openpyxl import load_workbook, Workbook


def cargar_excel(carpeta, fichero):
    #cargar el fichero de excel en nuestro archivo
    excel = load_workbook(f'./{carpeta}/{fichero}')
    # seleccionar la hoja que quiero. Hoja activa.
    hoja = excel.active
    # en caso de necesitar seleccionar una hoja concreta lo haremos de esta forma
    # hoja = excel['Finanzas'] siendo finanza el nombre de la hoja
    lista_empleados = []

    #extraer la primera fila del excel como cabeceras
    filas = hoja.iter_rows(values_only = True)
    cabeceras = next(filas)


    # recorrer filas y columnas de la hoja de excel

    # for fila in hoja.iter_rows(min_row=2, values_only=True):
    #     for value in fila:
    #         print(value)

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        empleado_dict = dict(zip(cabeceras, fila))
        lista_empleados.append(empleado_dict)
    return lista_empleados

lista = cargar_excel('data', 'empleados.xlsx')
print(lista)
